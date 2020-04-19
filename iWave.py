import requests
import openpyxl
from pprint import pprint

def getHeaders():
    headers = {
    'Accept': 'application/json',
    'Content-Type': 'application/json'
    }
    return headers

def getUser():
    url = 'https://pro.iwave.com/pro360#search:Overview_search:165659229'
    response = requests.get(url=url, headers=getHeaders(), auth=('glasere@purdue.edu', '5zqe4grI!!!!'))
    print(response.content)

#reorganizes the PROfile names into new columns - first name and last name
def excelOps(loc):
    wb = openpyxl.load_workbook(filename = loc)
    ws = wb.get_sheet_by_name("AmplifyOriginal")

    headers = {}
    row = 1
    column = 1
    header = ws.cell(row=row,column=column).value

    while header:
        headers[header] = column
        column += 1
        header = ws.cell(row=row,column=column).value

    firstColumn = headers['*First Name']
    lastColumn = headers['*Last Name']
    nameColumn = headers['PROfileName']
    row = 2
    name = ws.cell(row=row,column=nameColumn).value
    delim = ' '

    while name:
        first = delim.join(name.split(' ')[:-1])
        print(first)
        ws.cell(row=row,column=firstColumn).value = first
        ws.cell(row=row,column=lastColumn).value = name.split(' ')[-1]
        row += 1
        name = ws.cell(row=row,column=nameColumn).value



        #read the csv
        #enumerate the rows, so that you can
        #get the row index for the xlsx
            #Assuming space separated,
            #Split the row to cells (column)
            #row = row[0].split()
            #Access the particular cell and assign
            #the value from the csv row
    #save the csb file
    wb.save('AmplifyUpdated.xlsx')

def getColumnKey(ws):
    headers = {}
    row = 1
    column = 1
    header = ws.cell(row=row,column=column).value

    while header:
        headers[header] = column
        column += 1
        header = ws.cell(row=row,column=column).value
    
    return headers

def mergeInfo(loc1, loc2):
    wb = openpyxl.load_workbook(filename = loc1)
    ws = wb.get_sheet_by_name("Sheet1")
    wb2 = openpyxl.load_workbook(filename = loc2)
    ws2 = wb2.get_sheet_by_name("AmplifyOriginal")

    key = getColumnKey(ws)
    key2 = getColumnKey(ws2)

    donors = []
    firstColumn = key2['*First Name']
    row = 2
    name = ws2.cell(row=row,column=firstColumn).value

    while name:
        donors.append(donor(name, ws2.cell(row=row,column=key2['*Last Name']).value, ws2.cell(row=row,column=key2['Address1']).value, ws2.cell(row=row,column=key2['*City']).value, ws2.cell(row=row,column=key2['*State/Province']).value, ws2.cell(row=row,column=key2['ZIP/Postal Code']).value))
        row += 1
        name = ws2.cell(row=row,column=firstColumn).value


    firstColumn = key['First Name']
    lastColumn = key['Last Name']
    addressColumn = key['Street Address']
    cityColumn = key['City']
    stateColumn = key['State ']
    zipColumn = key['Zip']
    #phoneColumn = key['Phone']
    #emailColumn = key['Email address']

    row = 2
    name = ws.cell(row=row,column=firstColumn).value

    while name:
        for d in donors:
            if d.first == name and d.last == ws.cell(row=row,column=lastColumn).value:
                if d.address and len(d.address.split()) > 2 and d.state == 'Minnesota':
                    ws.cell(row=row,column=firstColumn).value = d.first
                    ws.cell(row=row,column=lastColumn).value = d.last
                    ws.cell(row=row,column=addressColumn).value = d.address
                    ws.cell(row=row,column=cityColumn).value = d.city
                    ws.cell(row=row,column=stateColumn).value = d.state
                    ws.cell(row=row,column=zipColumn).value = d.zip
        row += 1
        name = ws.cell(row=row,column=firstColumn).value

    wb.save('PostcardUpdated.xlsx')    


class donor:
    def __init__(self, first, last, address, city, state, zip):
        self.first = first
        self.last = last
        self.address = address
        self.city = city
        self.state = state
        self.zip = zip

if __name__ == "__main__":
    #getUser()
    # Give the location of the file 
    #loc = "../../Documents/AmplifyOriginal.xlsx"
    #excelOps(loc)
    loc1 = "../../Documents/PostcardOriginal.xlsx"
    loc2 = "AmplifyUpdated.xlsx"
    mergeInfo(loc1, loc2)